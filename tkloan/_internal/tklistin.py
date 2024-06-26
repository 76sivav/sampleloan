from openpyxl import load_workbook,worksheet,workbook,Workbook
import numpy as np
import os,sys
from docxtpl import DocxTemplate
from tkinter import *
from tkcalendar import *
from datetime import date,datetime

def resource(relative_path):
     try:
          base_path=sys._MEIPASS
     except Exception:
           base_path=os.path.abspath(".")
     return os.path.join(base_path,relative_path)

try:
    wb=load_workbook(resource("src/new.xlsx"))
    ws=wb.active
    if ws["A1"].value=="loan_date":
        pass
    else:
        head=["loan_date","bill_no","name","co_name","street","address","int_amt","weight","item","no_item","Phone No","release"]
        ws.append(head)
        wb.save(resource("src/new.xlsx"))
    
except:
    wb=Workbook()
    ws=wb.active
    head=["loan_date","bill_no","name","co_name","street","address","int_amt","weight","item","no_item","Phone No","release"]
    ws.append(head)
    wb.save(resource("src/new.xlsx")) 

def save(a):#for save not print or edit
    ws.append(a)
    wb.save(resource("src/new.xlsx")) 
    return True

def loanprint(a,c,l):
    if l=="loan":
        save(a)
    newloan=DocxTemplate(resource('src/loan.docx')) 
    bill=a[1]
    date=a[0]
    name=a[2]
    coname=a[3]
    street=a[4]
    address=a[5]
    item=a[8]
    weight=a[7]
    amount=a[6]
    noitem=a[9]
    phnum=a[10]
    maxd=datechange(date)
    m=datetime.strptime((maxd),'%d-%m-%Y')
    max_date=(m.replace(year=m.year+1))
    max_date=max_date.strftime('%d-%m-%Y')
    
    newloan.render({'loan_day':date,'bill_no':bill,'name':name,'coname':coname,'address':address,'street':street,'item':item,'weight':weight,'amount':amount,'noitem':noitem,'max_date':max_date})
    loan_name=f"{bill} {name} {l}.docx"
    if l=="loan":
        loan_path=r'.\loan'
    elif l=="reprint":
        loan_path=r'.\reprint'
    n_path=resource(os.path.join(loan_path,loan_name))
    newloan.save(n_path)
    if c==True:
        filepath=loan_path
        os.startfile(n_path,'print')


def interestprint(paylist,c):
    doc=DocxTemplate(resource('src/omm.docx'))  
    name=paylist[0]
    loan_date=paylist[1]
    bill_no=paylist[2]
    int_amt=paylist[3]
    to_day=paylist[4]
    interest=paylist[5]
    total=paylist[6]

    id=int(paylist[7])+1
    ws[f"l{id}"]=to_day
    wb.save(resource("src/new.xlsx"))

    doc.render({'name':name,'to_day':to_day,'loan_date':loan_date,'bill_no':bill_no,'interest':interest,'total':total,'int_amt':int_amt})
    r_name=f"{bill_no} {name}.docx"
    patth=r'.\relese'
    r_path=resource(os.path.join(patth,r_name))
    doc.save(r_path)
    if c==True:
        filepath=r_path
        os.startfile(filepath,'print')

def max_bill():
    s=list(ws.columns)[1]
    c=[]
    for i in range(1,ws.max_row):
        try:
            a=str(s[i].value)
            c.append(int(a))
        except:
            pass
    c.sort()
    return (c[-1]+1)



def alter(id,uplist):
        
        id+=1
        ws[f"C{id}"]=uplist[2]
        ws[f"D{id}"]=uplist[3]
        ws[f"A{id}"]=uplist[0]
        ws[f"B{id}"]=uplist[1]
        ws[f"F{id}"]=uplist[5]
        ws[f"H{id}"]=uplist[7]
        ws[f"I{id}"]=uplist[8]
        ws[f"G{id}"]=uplist[6]
        ws[f"E{id}"]=uplist[4]
        ws[f"J{id}"]=uplist[9]
        ws[f"k{id}"]=uplist[10]
        ws[f"l{id}"]=uplist[11]
        
        wb.save(resource("src/new.xlsx"))
    


def srch(i,val):
    
        id=val
        i=str(i)
        if i=="கடன் தேதி":
            src='loan_date'
        elif i=="கடன் எண்":
            src='bill_no'
        elif i=="பெயர்":
            src='name'
        elif i=="த/க பெயர்":
            src='co_name'
        elif i=="ஊர்":
            src='address'
        elif i=="கடன் தொகை":
            src='amount'
        elif i=="பொருள்":
            src='items'
        elif i=="மீட்ட தேதி":
            src='relese_date'
        elif i=="எடை":
            src='weight'
        elif i=='Phone No':
            src="Phone No"
        else:
            src="none"
        

        
        op=[]

        if src=='loan_date':
            id=datechange(id)
            s=list(ws.columns)[0]
            for i in range(1,(ws.max_row)):
                if datechange(s[i].value)==(id):
                    op.append(i)
    
        elif src=='bill_no':
            s=list(ws.columns)[1]
            
            for i in range(1,ws.max_row):
                a=str(s[i].value)
                try:
                    if int(a)==int(id):
                        op.append(i)
                except:
                    pass

        elif src=='name':
            s=np.array(list(ws.columns)[2])
            for i in range(1,ws.max_row):
                if str(id) in str(s[i].value):
                    op.append(i)

        elif src=='co_name':
            s=list(ws.columns)[3]
            for i in range(1,ws.max_row):
                if str(id) in str(s[i].value):
                    op.append(i)

        elif src=='address':
            s=np.array(list(ws.columns)[5])
            for i in range(1,ws.max_row):
                if str(id) in str(s[i].value):
                    op.append(i)

        elif src=='amount':
            s=list(ws.columns)[6]
            for i in range(1,ws.max_row):
                a=str(s[i].value)
                try:
                    if int(a)==int(id):
                        op.append(i)
                except:
                    pass

        elif src=='weight':
            s=list(ws.columns)[7]
            for i in range(1,ws.max_row):
                a=str(s[i].value)
                try:
                    if int(a)==int(id):
                        op.append(i)
                except:
                    pass
        
        elif src=='items':
            s=list(ws.columns)[8]
            for i in range(1,ws.max_row):
                if str(id) in str(s[i].value):
                    op.append(i)

        elif src=='Phone No':
            s=list(ws.columns)[10]
            for i in range(1,ws.max_row):
                if str(id) in str(s[i].value):
                    op.append(i)

        elif src=='relese_date':
            id=datechange(id)
            s=list(ws.columns)[11]
            for i in range(1,ws.max_row):
                if datechange(s[i].value)==datechange(id):
                    op.append(i)
        x=[]
        for a in op:
            c=list(ws.rows)[int(a)]
            v=[]
            for i in range(0,12):
                # if c[i].value==None:
                #     pass
                v.append(c[i].value)
            v.append(a)
            x.append(v)
        return x

def delete(id):
    ws.delete_rows((int(id)+1))
    wb.save(resource("src/new.xlsx"))

    
def datechange(a):        
        t1="%d-%m-%y %H:%M:%S"
        t2="%d-%m-%y"
        t3="%d %m %y %H:%M:%S"
        t4="%d %m %y"
        t5="%d/%m/%y %H:%M:%S"
        t6="%d/%m/%y"
        t7="%d-%m-%Y %H:%M:%S"
        t8="%d-%m-%Y"
        t9="%d %m %Y %H:%M:%S"
        t10="%d %m %Y"
        t11="%d/%m/%Y %H:%M:%S"
        t12="%d/%m/%Y"
        t13="%Y-%m-%d %H:%M:%S"
        t14="%Y-%m-%d"
        t=[t1,t2,t3,t4,t5,t6,t7,t8,t9,t10,t11,t12,t13,t14]
        try:
            for i in t:
                try:
                    a=datetime.strptime(a,i)
                    
                except:
                    pass
            return a.strftime("%d-%m-%Y")
        except:
            return a

def interest(int_amt,loan_date,re_date):
    loan_date=datetime.strptime(loan_date,"%d-%m-%Y")
    re_date=datetime.strptime(re_date,"%d-%m-%Y")
    diff=re_date-loan_date
    intday=int(diff.days)
    if intday<15:
        intday=15
    intrest=(int(int_amt)*0.015*intday)/30
    total=int(intrest)+int(int_amt)
    
    return (int(intrest))  
